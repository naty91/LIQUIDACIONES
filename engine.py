from __future__ import annotations

import io
import json
import math
from dataclasses import dataclass
from datetime import date, datetime
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook

SBU_2026 = 482.00
TOL_DEFAULT = 0.02

LEGAL = {
    "decimo_tercero": {
        "short": "Décimo tercero — Código del Trabajo, arts. 111 y 95.",
        "detail": "Corresponde a la doceava parte de las remuneraciones computables. El período de acumulación es del 1 de diciembre al 30 de noviembre cuando el trabajador pidió acumulación.",
        "url": "https://calculadoras.trabajo.gob.ec/tercero",
    },
    "decimo_cuarto": {
        "short": "Décimo cuarto — Código del Trabajo, art. 113; Acuerdo MDT-2023-140.",
        "detail": "Equivale a un SBU anual, proporcional al tiempo aplicable. Para acumulados: Costa/Insular, 1 de marzo a último día de febrero; Sierra/Amazonía, 1 de agosto a 31 de julio.",
        "url": "https://www.trabajo.gob.ec/29-cual-es-el-plazo-y-como-se-debe-realizar-la-solicitud-para-la-acumulacion-del-pago-de-la-decima-tercera-y-decima-cuarta-remuneracion/",
    },
    "vacaciones": {
        "short": "Vacaciones — Código del Trabajo, arts. 69, 71 y 76.",
        "detail": "El trabajador tiene 15 días anuales. Para la parte proporcional se usa la regla de la veinticuatroava parte de la remuneración computable del período; deben descontarse vacaciones ya gozadas/pagadas y revisar períodos acumulados.",
        "url": "https://www.trabajo.gob.ec/wp-content/uploads/downloads/2024/01/CODIGO_DEL_TRABAJO.pdf",
    },
    "desahucio": {
        "short": "Bonificación por desahucio — Código del Trabajo, arts. 184 y 185.",
        "detail": "25% de la última remuneración mensual por cada año de servicio en los casos en que legalmente corresponde. En despido intempestivo, el Ministerio indica considerarla si el trabajador cumplió un año o más.",
        "url": "https://www.trabajo.gob.ec/8-en-caso-de-despido-intempestivo-que-rubros-se-deben-tomar-en-cuenta-al-momento-de-realizar-la-liquidacion-del-trabajador/",
    },
    "despido": {
        "short": "Despido intempestivo — Código del Trabajo, art. 188.",
        "detail": "Hasta 3 años: 3 remuneraciones. Más de 3 años: 1 remuneración por cada año; la fracción se considera año completo; máximo 25 remuneraciones.",
        "url": "https://www.trabajo.gob.ec/8-en-caso-de-despido-intempestivo-que-rubros-se-deben-tomar-en-cuenta-al-momento-de-realizar-la-liquidacion-del-trabajador/",
    },
    "general": {
        "short": "Terminación laboral — Código del Trabajo, art. 169.",
        "detail": "La causal de terminación determina qué indemnizaciones o bonificaciones aplican, sin excluir los beneficios proporcionales y valores pendientes.",
        "url": "https://calculadoras.trabajo.gob.ec/liquidaciones",
    },
}

OFFICIAL_URLS = [
    "https://calculadoras.trabajo.gob.ec/liquidaciones",
    "https://calculadoras.trabajo.gob.ec/tercero",
    "https://www.trabajo.gob.ec/8-en-caso-de-despido-intempestivo-que-rubros-se-deben-tomar-en-cuenta-al-momento-de-realizar-la-liquidacion-del-trabajador/",
    "https://www.trabajo.gob.ec/29-cual-es-el-plazo-y-como-se-debe-realizar-la-solicitud-para-la-acumulacion-del-pago-de-la-decima-tercera-y-decima-cuarta-remuneracion/",
    "https://www.trabajo.gob.ec/wp-content/uploads/downloads/2024/01/CODIGO_DEL_TRABAJO.pdf",
]

CAUSALES = [
    "Desahucio solicitado por el trabajador",
    "Acuerdo entre las partes",
    "Despido intempestivo",
    "Visto bueno a favor del empleador",
    "Terminación por plazo / obra / servicio",
    "Liquidación del negocio con aviso previo",
    "Liquidación del negocio sin aviso previo",
    "Otra causal",
]


def money(v: float) -> str:
    try:
        n = float(v)
    except Exception:
        n = 0.0
    s = f"{n:,.2f}"
    # Ecuador visual format
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"$ {s}"


def num(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, bool):
        return float(v)
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(" ", "")
    if not s:
        return 0.0
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        # Excel 1900 date system
        try:
            return (datetime(1899, 12, 30) + pd.to_timedelta(float(v), unit="D")).date()
        except Exception:
            return None
    if isinstance(v, str) and v.strip():
        for dayfirst in (True, False):
            try:
                return pd.to_datetime(v, dayfirst=dayfirst).date()
            except Exception:
                pass
    return None


def days360_us(start: date, end: date) -> int:
    """Approximate Excel DAYS360 US/NASD, sufficient for payroll proportional controls."""
    if not start or not end or end < start:
        return 0
    d1 = start.day
    d2 = end.day
    m1 = start.month
    m2 = end.month
    y1 = start.year
    y2 = end.year
    if d1 == 31:
        d1 = 30
    if d2 == 31 and d1 >= 30:
        d2 = 30
    return (y2 - y1) * 360 + (m2 - m1) * 30 + (d2 - d1)


def completed_years(start: date, end: date) -> int:
    if not start or not end or end < start:
        return 0
    return max(0, end.year - start.year - ((end.month, end.day) < (start.month, start.day)))


def years_for_dismissal(start: date, end: date) -> int:
    if not start or not end or end < start:
        return 0
    y = completed_years(start, end)
    anniversary = date(start.year + y, start.month, min(start.day, 28 if start.month == 2 else start.day))
    has_fraction = end > anniversary
    return y + (1 if has_fraction else 0)


def default_desahucio(causal: str, start: date, end: date) -> bool:
    yrs = completed_years(start, end)
    if causal in {"Desahucio solicitado por el trabajador", "Acuerdo entre las partes", "Liquidación del negocio con aviso previo", "Liquidación del negocio sin aviso previo"}:
        return yrs >= 1
    if causal == "Despido intempestivo":
        return yrs >= 1
    return False


def default_despido(causal: str) -> bool:
    return causal in {"Despido intempestivo", "Liquidación del negocio sin aviso previo"}


def dec14_period(end: date, region: str):
    if region.startswith("Costa"):
        if end.month >= 3:
            return date(end.year, 3, 1), date(end.year + 1, 2, 28)
        return date(end.year - 1, 3, 1), date(end.year, 2, 28)
    # Sierra / Amazonía: Aug-Jul
    if end.month >= 8:
        return date(end.year, 8, 1), date(end.year + 1, 7, 31)
    return date(end.year - 1, 8, 1), date(end.year, 7, 31)


def dec13_period(end: date):
    if end.month == 12:
        return date(end.year, 12, 1), date(end.year + 1, 11, 30)
    return date(end.year - 1, 12, 1), date(end.year, 11, 30)


def calc_dec14_days(start: date, end: date, region: str) -> int:
    p_start, p_end = dec14_period(end, region)
    a = max(start, p_start)
    b = min(end, p_end)
    if b < a:
        return 0
    return min(360, days360_us(a, b) + 1)


def normalize_detail(df: pd.DataFrame) -> pd.DataFrame:
    wanted = ["Mes", "Remuneración computable", "Días"]
    if df is None or df.empty:
        return pd.DataFrame(columns=wanted)
    out = df.copy()
    for col in wanted:
        if col not in out.columns:
            out[col] = "" if col == "Mes" else 0.0
    out = out[wanted]
    out["Mes"] = out["Mes"].fillna("").astype(str)
    out["Remuneración computable"] = out["Remuneración computable"].map(num)
    out["Días"] = out["Días"].map(num)
    out = out[(out["Mes"].str.strip() != "") | (out["Remuneración computable"] != 0) | (out["Días"] != 0)].reset_index(drop=True)
    return out


@dataclass
class VerificationInput:
    name: str
    ident: str
    start: date
    end: date
    region: str
    causal: str
    last_salary: float
    sbu: float
    detail: pd.DataFrame
    dec13_accumulated: bool
    dec14_accumulated: bool
    vacation_base_override: float
    vacation_adjustment: float
    apply_desahucio: bool
    apply_dismissal: bool
    pending_salary_expected: float
    reserve_fund_expected: float
    other_expected: float
    reported: dict
    tolerance: float = TOL_DEFAULT


def verify(inp: VerificationInput):
    detail = normalize_detail(inp.detail)
    total_base = float(detail["Remuneración computable"].sum())
    total_days = float(detail["Días"].sum())

    # The CENASE format uses the monthly computable remuneration base for both D13 and proportional vacation.
    calc13 = total_base / 12.0 if inp.dec13_accumulated else 0.0

    d14_days = calc_dec14_days(inp.start, inp.end, inp.region) if inp.dec14_accumulated else 0
    calc14 = (inp.sbu / 360.0) * d14_days if inp.dec14_accumulated else 0.0

    vac_base = inp.vacation_base_override if inp.vacation_base_override > 0 else total_base
    calc_vac = max(0.0, vac_base / 24.0 + inp.vacation_adjustment)

    yrs = completed_years(inp.start, inp.end)
    calc_des = inp.last_salary * 0.25 * yrs if inp.apply_desahucio and yrs >= 1 else 0.0

    yd = years_for_dismissal(inp.start, inp.end)
    if inp.apply_dismissal and yd > 0:
        months = 3 if yd <= 3 else min(yd, 25)
        calc_dismiss = inp.last_salary * months
    else:
        months = 0
        calc_dismiss = 0.0

    expected = [
        ("Décimo tercero", calc13, "decimo_tercero"),
        ("Décimo cuarto", calc14, "decimo_cuarto"),
        ("Vacaciones", calc_vac, "vacaciones"),
        ("Bonificación por desahucio", calc_des, "desahucio"),
        ("Indemnización por despido intempestivo", calc_dismiss, "despido"),
        ("Remuneración / sueldo pendiente", inp.pending_salary_expected, "general"),
        ("Fondos de reserva pendientes", inp.reserve_fund_expected, "general"),
        ("Otros valores a favor del trabajador", inp.other_expected, "general"),
    ]

    rows = []
    for concept, calc, key in expected:
        rep = num(inp.reported.get(concept, 0.0))
        calc = round(float(calc), 2)
        rep = round(rep, 2)
        diff = round(rep - calc, 2)
        if abs(diff) <= inp.tolerance:
            status = "✅ CORRECTO"
        elif calc > 0 and rep == 0:
            status = "❌ RUBRO OMITIDO"
        elif calc == 0 and rep > inp.tolerance:
            status = "⚠️ REVISAR APLICACIÓN"
        else:
            status = "⚠️ DIFERENCIA"
        rows.append({
            "Concepto": concept,
            "Calculado por APP": calc,
            "Reportado RR.HH.": rep,
            "Diferencia (RR.HH.-APP)": diff,
            "Resultado": status,
            "Base legal corta": LEGAL[key]["short"],
            "Fuente oficial": LEGAL[key]["url"],
        })

    summary = pd.DataFrame(rows)
    material = summary[~((summary["Calculado por APP"] == 0) & (summary["Reportado RR.HH."] == 0))]
    failures = int((material["Resultado"] != "✅ CORRECTO").sum())
    verdict = "✅ APTO PARA PAGO" if failures == 0 else f"⚠️ REQUIERE CORRECCIÓN / REVISIÓN ({failures} observación(es))"

    meta = {
        "base_d13": round(total_base, 2),
        "base_vac": round(vac_base, 2),
        "days_detail": round(total_days, 2),
        "d14_days": d14_days,
        "years_completed": yrs,
        "dismissal_years": yd,
        "dismissal_months": months,
        "total_calc": round(float(summary["Calculado por APP"].sum()), 2),
        "total_reported": round(float(summary["Reportado RR.HH."].sum()), 2),
        "total_diff": round(float(summary["Reportado RR.HH."].sum() - summary["Calculado por APP"].sum()), 2),
        "verdict": verdict,
        "failures": failures,
    }
    return summary, meta, detail


def extract_cenase_excel(file_obj):
    """Read the current CENASE VERIFICADOR layout without modifying the workbook."""
    wb = load_workbook(file_obj, data_only=True, read_only=True, keep_vba=False)
    if "VERIFICADOR" not in wb.sheetnames:
        raise ValueError("No se encontró la hoja 'VERIFICADOR'.")
    ws = wb["VERIFICADOR"]

    name = str(ws["A1"].value or "").strip()
    start = as_date(ws["A2"].value)
    end = as_date(ws["A3"].value)

    rows = []
    receive_row = None
    for r in range(5, min(ws.max_row, 120) + 1):
        a, b, c, d = [ws.cell(r, col).value for col in range(1, 5)]
        label_a = str(a or "").strip().lower()
        if "a recibir" in label_a:
            receive_row = r
            break
        # Detail rows in the CENASE format always have a month/date in column A.
        # This intentionally excludes the totals row, which has column A blank.
        if a not in (None, "") and label_a not in {"mes"} and num(b) != 0:
            rows.append({
                "Mes": a.strftime("%b-%y") if isinstance(a, (datetime, date)) else str(a or ""),
                "Remuneración computable": num(b),
                "Días": num(c),
            })

    reported13 = num(ws.cell(receive_row, 2).value) if receive_row else 0.0
    reported_vac = num(ws.cell(receive_row, 4).value) if receive_row else 0.0
    reported_des = 0.0
    total_reported = 0.0
    notes = []

    # Read only the active result block immediately below “a recibir:”.
    # The workbook contains other scenario/calculation blocks further down that must not be mixed in.
    if receive_row:
        for r in range(receive_row + 1, min(ws.max_row, receive_row + 20) + 1):
            btxt = str(ws.cell(r, 2).value or "").strip().lower()
            dval = num(ws.cell(r, 4).value)
            if btxt.startswith("desahucio") or btxt.startswith("desahusio"):
                reported_des = dval
            if "total a recibir" in btxt:
                total_reported = dval
                break

    # Cédula often exists in the CONFIRMADOR sheet for selected worker, but layout can vary.
    ident = ""
    if "CONFIRMADOR" in wb.sheetnames:
        wc = wb["CONFIRMADOR"]
        for r in range(1, min(wc.max_row, 50) + 1):
            for c in range(1, min(wc.max_column, 8) + 1):
                v = wc.cell(r, c).value
                if isinstance(v, str) and v.isdigit() and len(v) == 10:
                    ident = v
                    break
            if ident:
                break

    if receive_row is None:
        notes.append("No se detectó la fila 'a recibir:'; revise los valores reportados manualmente.")

    return {
        "name": name,
        "ident": ident,
        "start": start,
        "end": end,
        "rows": rows,
        "reported13": reported13,
        "reported_vac": reported_vac,
        "reported_des": reported_des,
        "reported_total": total_reported,
        "notes": notes,
    }


def backup_payload(data: dict) -> bytes:
    def conv(o):
        if isinstance(o, (date, datetime)):
            return o.isoformat()
        if isinstance(o, pd.DataFrame):
            return o.to_dict("records")
        raise TypeError(type(o).__name__)
    return json.dumps(data, ensure_ascii=False, indent=2, default=conv).encode("utf-8")


def restore_payload(raw: bytes) -> dict:
    return json.loads(raw.decode("utf-8"))


def _looks_name(v):
    if not isinstance(v, str):
        return False
    s=v.strip()
    if not s or s.lower() in {'mes','a recibir:','a recibir'}:
        return False
    return any(ch.isalpha() for ch in s) and len(s) >= 5


def _is_date(v):
    return isinstance(v, (datetime, date))


def _month_num(v):
    """Return month number from CENASE month label/date cells."""
    if isinstance(v, (datetime, date)):
        return v.month
    txt = str(v or "").strip().lower()
    if not txt:
        return None
    months = {
        "ene": 1, "jan": 1, "feb": 2, "mar": 3, "abr": 4, "apr": 4,
        "may": 5, "jun": 6, "jul": 7, "ago": 8, "aug": 8,
        "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dic": 12, "dec": 12,
    }
    low = txt[:4].replace('.', '')
    for k, m in months.items():
        if txt.startswith(k) or low.startswith(k):
            return m
    return None


def _assign_years_to_months(months, end: date):
    """Infer calendar years for sequential month labels, working backwards from exit year."""
    if not months:
        return []
    years = [None] * len(months)
    y = end.year if end else date.today().year
    next_m = None
    for i in range(len(months) - 1, -1, -1):
        m = months[i]
        if m is None:
            years[i] = y
            continue
        if next_m is not None and m > next_m:
            y -= 1
        years[i] = y
        next_m = m
    return years


def _payroll_days_for_month(start: date, end: date, year: int, month: int) -> int:
    """Expected payroll days on CENASE's 30-day monthly convention.

    Full month = 30 days. Entry day and exit day are inclusive; day 31 is treated as day 30.
    Examples: entry Jan-03 => 28 days in Jan; entry Jun-25 => 6 days in Jun;
    exit Aug-27 => 27 days in Aug; full February => 30 days.
    """
    if not start or not end or end < start or not year or not month:
        return 0
    month_start = date(year, month, 1)
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    # If employment does not overlap this calendar month.
    if end < month_start or start >= next_month:
        return 0
    first = 1
    last = 30
    if start.year == year and start.month == month:
        first = min(start.day, 30)
    if end.year == year and end.month == month:
        last = min(end.day, 30)
    return max(0, last - first + 1)


def extract_cenase_batch(file_obj):
    """Extract and audit all consecutive CENASE liquidation blocks from a workbook.

    In addition to benefits, this version validates the DAYS column against entry/exit dates
    using the company's 30-day payroll convention, month by month and in total.
    """
    wb = load_workbook(file_obj, data_only=True, read_only=True, keep_vba=False)
    out = []
    for ws in wb.worksheets:
        maxr = ws.max_row
        r = 1
        while r <= maxr - 3:
            a = ws.cell(r, 1).value
            if _looks_name(a) and _is_date(ws.cell(r + 1, 1).value) and _is_date(ws.cell(r + 2, 1).value) and str(ws.cell(r + 3, 1).value or '').strip().lower() == 'mes':
                start_row = r
                name = str(a).strip()
                # La cédula suele estar a la derecha del nombre en el formato masivo.
                ident = ''
                for cc in range(2, min(ws.max_column, 10) + 1):
                    cand = normalize_id(ws.cell(r, cc).value)
                    if len(cand) == 10:
                        ident = cand
                        break
                start = as_date(ws.cell(r + 1, 1).value)
                end = as_date(ws.cell(r + 2, 1).value)
                rows = []
                j = r + 4
                receive = None
                reported_total_days = None
                while j <= maxr:
                    c0, c1, c2, c3 = [ws.cell(j, c).value for c in range(1, 5)]
                    la = str(c0 or '').strip().lower()
                    if 'a recibir' in la:
                        receive = j
                        break
                    if c0 not in (None, '') and (num(c1) != 0 or num(c2) != 0 or num(c3) != 0):
                        rows.append({
                            'Mes': c0.strftime('%b') if _is_date(c0) else str(c0),
                            '_month': _month_num(c0),
                            'Remuneración computable': num(c1),
                            'Días': num(c2),
                        })
                    elif c0 in (None, '') and (num(c1) != 0 or num(c2) != 0 or num(c3) != 0):
                        reported_total_days = num(c2)
                    j += 1

                # Day audit: assign a real year to each month label and calculate what should apply.
                month_nums = [x.get('_month') for x in rows]
                years = _assign_years_to_months(month_nums, end)
                day_checks = []
                for idx, row in enumerate(rows):
                    m = month_nums[idx]
                    y = years[idx]
                    expected = _payroll_days_for_month(start, end, y, m) if m else 0
                    reported = round(num(row.get('Días')), 2)
                    diff = round(reported - expected, 2)
                    ok_days = abs(diff) <= 0.01
                    row['Año inferido'] = y
                    row['Días según fechas'] = expected
                    row['Diferencia días'] = diff
                    row['Estado días'] = '✅ CORRECTO' if ok_days else '⚠️ REVISAR'
                    row.pop('_month', None)
                    day_checks.append({
                        'Mes': row['Mes'], 'Año': y, 'Días RR.HH.': reported,
                        'Días según fechas': expected, 'Diferencia': diff,
                        'Estado': '✅ CORRECTO' if ok_days else '⚠️ REVISAR'
                    })

                calculated_days = round(sum(x['Días según fechas'] for x in rows), 2)
                detail_days = round(sum(num(x['Días']) for x in rows), 2)
                if reported_total_days is None:
                    reported_total_days = detail_days
                reported_total_days = round(num(reported_total_days), 2)
                total_days_ok = abs(reported_total_days - calculated_days) <= 0.01 and all(x['Estado'] == '✅ CORRECTO' for x in day_checks)

                rep13 = repvac = repdes = reptotal = 0.0
                des_years = 0
                if receive:
                    rep13 = num(ws.cell(receive, 2).value)
                    repvac = num(ws.cell(receive, 4).value)
                    for k in range(receive + 1, min(maxr, receive + 10) + 1):
                        txt = str(ws.cell(k, 2).value or '').strip()
                        low = txt.lower()
                        val = num(ws.cell(k, 4).value)
                        if low.startswith('desahucio') or low.startswith('desahusio'):
                            repdes = val
                            import re
                            m = re.search(r'(\d+)\s*a', low)
                            if m:
                                des_years = int(m.group(1))
                        if 'total a recibir' in low:
                            reptotal = val
                            break

                last_salary = rows[-1]['Remuneración computable'] if rows else 0.0
                base = round(sum(x['Remuneración computable'] for x in rows), 2)
                calc13 = round(base / 12, 2)
                calcvac = round(base / 24, 2)
                calcdes = round(last_salary * 0.25 * des_years, 2) if des_years else 0.0
                calctotal = round(calcvac + calcdes, 2)
                checks = []
                def ok(rep, calc):
                    return abs(float(rep) - float(calc)) <= TOL_DEFAULT
                checks.append(('Décimo tercero', rep13, calc13, ok(rep13, calc13)))
                checks.append(('Vacaciones', repvac, calcvac, ok(repvac, calcvac)))
                if des_years or repdes:
                    checks.append(('Desahucio', repdes, calcdes, ok(repdes, calcdes)))
                checks.append(('Total a recibir', reptotal, calctotal, ok(reptotal, calctotal)))
                money_failures = sum(1 for _, _, _, flag in checks if not flag)
                day_failures = sum(1 for x in day_checks if x['Estado'] != '✅ CORRECTO')
                if not total_days_ok and day_failures == 0:
                    day_failures = 1
                failures = money_failures + day_failures

                out.append({
                    'sheet': ws.title, 'row': start_row, 'name': name, 'ident': ident, 'start': start, 'end': end,
                    'rows': rows, 'day_checks': day_checks,
                    'base': base, 'days': detail_days, 'reported_total_days': reported_total_days,
                    'expected_days': calculated_days, 'days_diff': round(reported_total_days - calculated_days, 2),
                    'days_ok': total_days_ok, 'day_failures': day_failures,
                    'last_salary': last_salary,
                    'reported13': rep13, 'calc13': calc13, 'reported_vac': repvac, 'calc_vac': calcvac,
                    'des_years': des_years, 'reported_des': repdes, 'calc_des': calcdes,
                    'reported_total': reptotal, 'calc_total': calctotal, 'checks': checks,
                    'failures': failures,
                    'status': '✅ CORRECTO' if failures == 0 else f'⚠️ REVISAR ({failures})'
                })
                r = max(j + 1, r + 4)
            else:
                r += 1
    return out

# ===================== V5.1: CRUCES RRHH / APP / IESS / BASE PERSONAL =====================
import unicodedata
import re

IESS_PERSONAL_RATE = 0.0945
IESS_EMPLOYER_RATE = 0.1115
IESS_TOTAL_RATE = 0.2060


def normalize_text(v):
    s = str(v or '').strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^A-Z0-9 ]+', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def normalize_id(v):
    s = re.sub(r'\D+', '', str(v or ''))
    return s[-10:] if len(s) >= 10 else s


def employee_key(ident='', name=''):
    ident = normalize_id(ident)
    return f'ID:{ident}' if len(ident) == 10 else f'NM:{normalize_text(name)}'


def salary_for_days(monthly_salary: float, days: float) -> float:
    """Sueldo fijo proporcional usando divisor 30."""
    return round(num(monthly_salary) / 30.0 * num(days), 2) if num(monthly_salary) > 0 else 0.0


def iess_contributions(base: float):
    b = round(num(base), 2)
    return {
        'base': b,
        'personal': round(b * IESS_PERSONAL_RATE, 2),
        'patronal': round(b * IESS_EMPLOYER_RATE, 2),
        'total': round(b * IESS_TOTAL_RATE, 2),
    }


def read_tabular_excel(file_obj, sheet_name=None):
    """Lee una hoja de Excel manteniendo las columnas para mapeo flexible."""
    data = pd.read_excel(file_obj, sheet_name=sheet_name if sheet_name is not None else 0, dtype=object)
    data.columns = [str(c).strip() for c in data.columns]
    return data


def make_reference_index(df: pd.DataFrame, id_col=None, name_col=None):
    idx = {}
    if df is None or df.empty:
        return idx
    for _, row in df.iterrows():
        ident = row.get(id_col, '') if id_col else ''
        name = row.get(name_col, '') if name_col else ''
        key = employee_key(ident, name)
        if key not in {'ID:', 'NM:'}:
            idx.setdefault(key, []).append(row.to_dict())
        if name_col and normalize_text(name):
            idx.setdefault(f'NM:{normalize_text(name)}', []).append(row.to_dict())
    return idx


def lookup_reference(index, ident='', name=''):
    k_id = employee_key(ident, '') if normalize_id(ident) else None
    if k_id and k_id in index:
        return index[k_id][0], 'CÉDULA'
    k_nm = f'NM:{normalize_text(name)}'
    if normalize_text(name) and k_nm in index:
        return index[k_nm][0], 'NOMBRE'
    return None, ''


def parse_month_period(v):
    if v is None or v == '':
        return None
    if isinstance(v, (date, datetime)):
        return (v.year, v.month)
    try:
        d = pd.to_datetime(v, dayfirst=True, errors='coerce')
        if pd.notna(d):
            return (int(d.year), int(d.month))
    except Exception:
        pass
    txt = str(v).strip().lower()
    m = re.search(r'(20\d{2})\D{0,3}(0?[1-9]|1[0-2])', txt)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.search(r'(0?[1-9]|1[0-2])\D{0,3}(20\d{2})', txt)
    if m:
        return int(m.group(2)), int(m.group(1))
    mn = _month_num(txt)
    return (None, mn) if mn else None


def build_cross_check(liq_item, personnel_row=None, personnel_map=None, iess_rows=None, iess_map=None):
    """Tres capas: RRHH vs APP, APP vs IESS, y datos vs Base de Personal."""
    personnel_map = personnel_map or {}
    iess_map = iess_map or {}
    iess_rows = iess_rows or []

    real_start = liq_item.get('start')
    basic_salary = 0.0
    personnel_name = ''
    personnel_id = ''
    if personnel_row:
        personnel_name = str(personnel_row.get(personnel_map.get('name'), '') or '') if personnel_map.get('name') else ''
        personnel_id = normalize_id(personnel_row.get(personnel_map.get('id'), '')) if personnel_map.get('id') else ''
        if personnel_map.get('start'):
            real_start = as_date(personnel_row.get(personnel_map['start'])) or real_start
        if personnel_map.get('salary'):
            basic_salary = num(personnel_row.get(personnel_map['salary']))

    date_diff = None
    if real_start and liq_item.get('start'):
        date_diff = (liq_item['start'] - real_start).days

    real_day_checks = []
    for row in liq_item.get('rows', []):
        m = _month_num(row.get('Mes'))
        y = int(row.get('Año inferido') or (liq_item.get('end').year if liq_item.get('end') else date.today().year))
        expected = _payroll_days_for_month(real_start, liq_item.get('end'), y, m) if m else 0
        reported = num(row.get('Días'))
        fixed_salary_expected = salary_for_days(basic_salary, expected) if basic_salary else 0.0
        real_day_checks.append({
            'Mes': row.get('Mes'), 'Año': y, 'Días RR.HH.': reported,
            'Días APP según base personal': expected,
            'Dif. días': round(reported - expected, 2),
            'Salario básico mensual': round(basic_salary, 2),
            'Sueldo proporcional APP': fixed_salary_expected,
            'Estado días': '✅ CORRECTO' if abs(reported-expected) <= 0.01 else '⚠️ REVISAR',
        })

    iess_checks = []
    for rr in iess_rows:
        period = parse_month_period(rr.get(iess_map.get('period'))) if iess_map.get('period') else None
        i_days = num(rr.get(iess_map.get('days'))) if iess_map.get('days') else 0.0
        i_base = num(rr.get(iess_map.get('base'))) if iess_map.get('base') else 0.0
        i_personal = num(rr.get(iess_map.get('personal'))) if iess_map.get('personal') else 0.0
        i_patronal = num(rr.get(iess_map.get('patronal'))) if iess_map.get('patronal') else 0.0

        expected_days = 0
        expected_fixed = 0.0
        if period and period[1]:
            py = period[0] or (liq_item.get('end').year if liq_item.get('end') else date.today().year)
            expected_days = _payroll_days_for_month(real_start, liq_item.get('end'), py, period[1])
            expected_fixed = salary_for_days(basic_salary, expected_days) if basic_salary else 0.0

        contrib_base = i_base if i_base > 0 else expected_fixed
        calc_iess = iess_contributions(contrib_base)
        iess_checks.append({
            'Período': f'{period[1]:02d}/{period[0]}' if period and period[0] else str(rr.get(iess_map.get('period'), '') if iess_map.get('period') else ''),
            'Días IESS': i_days,
            'Días APP': expected_days,
            'Dif. días IESS-APP': round(i_days - expected_days, 2) if expected_days else '',
            'Materia gravada IESS': round(i_base, 2),
            'Sueldo proporcional básico APP': round(expected_fixed, 2),
            'Aporte personal IESS': round(i_personal, 2),
            'Aporte personal APP 9,45%': calc_iess['personal'],
            'Dif. personal': round(i_personal - calc_iess['personal'], 2) if i_personal else '',
            'Aporte patronal IESS': round(i_patronal, 2),
            'Aporte patronal APP 11,15%': calc_iess['patronal'],
            'Dif. patronal': round(i_patronal - calc_iess['patronal'], 2) if i_patronal else '',
            'Total aporte APP 20,60%': calc_iess['total'],
        })

    return {
        'personnel_name': personnel_name,
        'personnel_id': personnel_id,
        'real_start': real_start,
        'basic_salary': round(basic_salary, 2),
        'start_date_diff_days': date_diff,
        'day_checks_real': real_day_checks,
        'iess_checks': iess_checks,
    }

# ===================== V6: BDD CENASE AUTOMATICA (ACTIVOS / REINGRESOS / INACTIVOS) =====================
def load_cenase_personnel_database(file_obj, default_salary=SBU_2026):
    """Lee directamente la BDD CENASE y normaliza relaciones laborales históricas.

    Fuentes reconocidas: ACTIVOS (encabezado fila 5), REINGRESOS (fila 2),
    INACTIVOS (fila 5). La BDD actual no contiene sueldo básico; por ello se usa
    el salario básico de control indicado por el usuario (por defecto SBU 2026 = 482).
    """
    file_obj.seek(0)
    xls = pd.ExcelFile(file_obj)
    required = {'ACTIVOS','REINGRESOS','INACTIVOS'}
    if not required.issubset(set(xls.sheet_names)):
        raise ValueError('La BDD no contiene las hojas ACTIVOS, REINGRESOS e INACTIVOS esperadas.')
    records = []

    def add(ident, name, start, end=None, status='', source='', motive='', salary=default_salary, cycle=''):
        ident_n = normalize_id(ident); name_n = str(name or '').strip(); s=as_date(start); e=as_date(end)
        if not s or (not ident_n and not name_n): return
        records.append({'ident':ident_n,'name':name_n,'start':s,'end':e,'status':status,'source':source,
                        'motive':str(motive or '').strip(),'salary':num(salary) or num(default_salary),'cycle':cycle})

    file_obj.seek(0)
    act = pd.read_excel(file_obj, sheet_name='ACTIVOS', header=4, dtype=object)
    for _, r in act.iterrows():
        add(r.get('C. Identidad'), r.get('Nombres Completos'), r.get('F. INGRESO'), None,
            'ACTIVO','ACTIVOS','',default_salary,'VIGENTE')

    file_obj.seek(0)
    ina = pd.read_excel(file_obj, sheet_name='INACTIVOS', header=4, dtype=object)
    for _, r in ina.iterrows():
        add(r.get('C. Identidad'), r.get('Nombres Completos'), r.get('F. INGRESO'), r.get('F. SALIDA'),
            'INACTIVO','INACTIVOS',r.get('MOTIVO DE SALIDA'),default_salary,'HISTÓRICO')

    file_obj.seek(0)
    rei = pd.read_excel(file_obj, sheet_name='REINGRESOS', header=1, dtype=object)
    # Pandas desambigua columnas repetidas como F. INGRESO.1, F. INGRESO.2, etc.
    starts=[c for c in rei.columns if normalize_text(c).startswith('F INGRESO')]
    exits=[c for c in rei.columns if normalize_text(c).startswith('F SALIDA')]
    motives=[c for c in rei.columns if normalize_text(c).startswith('MOTIVO SALIDA')]
    for _, r in rei.iterrows():
        for i, sc in enumerate(starts):
            ec = exits[i] if i < len(exits) else None; mc = motives[i] if i < len(motives) else None
            add(r.get('C. Identidad'), r.get('Nombres Completos'), r.get(sc), r.get(ec) if ec else None,
                'REINGRESO','REINGRESOS',r.get(mc) if mc else '',default_salary,f'CICLO {i+1}')
    return records


def resolve_personnel_record(records, ident='', name='', liquidation_end=None):
    """Escoge el ciclo laboral que corresponde a la liquidación, no el primer ingreso histórico."""
    ident_n=normalize_id(ident); name_n=normalize_text(name); end=as_date(liquidation_end)
    candidates=[]
    for r in records or []:
        match = (ident_n and r['ident']==ident_n) or (name_n and normalize_text(r['name'])==name_n)
        if not match: continue
        if end and r['start'] > end: continue
        # Puntaje: cédula > nombre; salida exacta/cercana > ciclo abierto; ingreso más reciente.
        score = 1000 if ident_n and r['ident']==ident_n else 500
        if end and r.get('end'):
            delta=abs((r['end']-end).days)
            score += max(0, 300-delta)
            if r['end']==end: score += 500
        elif end and not r.get('end'):
            score += 150
        score += r['start'].toordinal()/1000000.0
        candidates.append((score,r))
    if not candidates: return None,''
    candidates.sort(key=lambda x:x[0], reverse=True)
    r=candidates[0][1]
    return r, ('CÉDULA' if ident_n and r['ident']==ident_n else 'NOMBRE')

# ===================== V7: IESS AUTOMÁTICO + BENEFICIOS POR PERÍODOS LEGALES =====================

def load_iess_cenase(file_obj):
    """Lee el formato real IESS BASE.xlsx de CENASE sin mapeo manual."""
    if hasattr(file_obj, 'seek'):
        file_obj.seek(0)
    df = pd.read_excel(file_obj, sheet_name='IESS', header=1, dtype=object)
    df.columns = [str(c).strip() for c in df.columns]
    required = ['Periodo','Cédula','Nombre','Rel. Trabajo','Sueldo','Días','Patronal','Individual']
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError('Faltan columnas esperadas en IESS: ' + ', '.join(missing))
    df = df[df['Cédula'].notna() | df['Nombre'].notna()].copy()
    df['Cédula_norm'] = df['Cédula'].map(normalize_id)
    df['Nombre_norm'] = df['Nombre'].map(normalize_text)
    df['_period'] = df['Periodo'].map(parse_month_period)
    return df


def iess_rows_for_employee(df, ident='', name=''):
    if df is None or df.empty:
        return []
    identn = normalize_id(ident)
    namen = normalize_text(name)
    sub = pd.DataFrame()
    if identn:
        sub = df[df['Cédula_norm'] == identn]
    if sub.empty and namen:
        sub = df[df['Nombre_norm'] == namen]
    return sub.to_dict('records') if not sub.empty else []


def _period_date(period):
    if not period or not period[0] or not period[1]:
        return None
    return date(int(period[0]), int(period[1]), 1)


def _period_in_range(period, start, end):
    d = _period_date(period)
    if not d or not start or not end:
        return False
    return (d.year, d.month) >= (start.year, start.month) and (d.year, d.month) <= (end.year, end.month)


def latest_iess_salary(irows, on_or_before=None):
    usable=[]
    for r in irows or []:
        p = r.get('_period') or parse_month_period(r.get('Periodo'))
        d = _period_date(p)
        if d and (on_or_before is None or (d.year,d.month) <= (on_or_before.year,on_or_before.month)):
            usable.append((d, num(r.get('Sueldo'))))
    usable.sort(key=lambda x:x[0])
    return usable[-1][1] if usable else 0.0


def adjusted_iess_base_for_days(base, iess_days, max_days_by_dates):
    """Devuelve la base IESS utilizable sin inflar meses realmente trabajados por menos días.

    Regla CENASE:
    - Un mes completo tiene tope 30 días (febrero 28/29 y meses de 31 se tratan como 30).
    - La fecha de ingreso BDD y la fecha de salida RR.HH. fijan el MÁXIMO de días posibles.
    - Si IESS reporta MENOS días que ese máximo (faltas, ingreso parcial u otra novedad),
      se conserva íntegramente la base IESS: nunca se aumenta artificialmente.
    - Solo si IESS reporta MÁS días que los permitidos por la fecha de salida/ingreso,
      se reduce proporcionalmente la base al máximo permitido.

    Ejemplos:
      IESS $238,22 / 14 días, máximo por fechas 28 -> se conserva $238,22.
      IESS $290,13 / 17 días, máximo por fechas 30 -> se conserva $290,13.
      IESS $482 / 30 días, salida RR.HH. día 27 -> $482/30*27 = $433,80.
    """
    base = num(base)
    iess_days = max(0.0, min(30.0, num(iess_days)))
    max_days = max(0.0, min(30.0, num(max_days_by_dates)))
    if base == 0 or max_days == 0:
        return 0.0
    # Nunca escalar hacia arriba una base IESS por días efectivamente inferiores.
    if iess_days <= 0 or iess_days <= max_days:
        return round(base, 2)
    # Solo recortar cuando IESS excede los días máximos permitidos por las fechas laborales.
    return round((base / iess_days) * max_days, 2)


def sum_iess_base(irows, start, end):
    """Suma bases IESS respetando días reales y recortando únicamente excesos por fechas.

    La fecha de ingreso BDD y la salida RR.HH. fijan el máximo de días posibles de cada mes.
    Si IESS tiene menos días, su base se conserva; si tiene más que el máximo, se recorta.
    """
    total=0.0
    detail=[]
    for r in irows or []:
        p = r.get('_period') or parse_month_period(r.get('Periodo'))
        if _period_in_range(p, start, end):
            base=num(r.get('Sueldo'))
            days=num(r.get('Días'))
            y = p[0] if p else None
            m = p[1] if p else None
            correct_days = _payroll_days_for_month(start, end, y, m) if y and m else days
            adj_base = adjusted_iess_base_for_days(base, days, correct_days)
            total += adj_base
            detail.append({
                'Periodo':r.get('Periodo'),'Año':y,'Mes':m,
                'Días IESS':days,'Días máximos por fechas':correct_days,
                'Base IESS reportada':round(base,2),'Base IESS utilizable':adj_base,
                'Ajuste':round(adj_base-base,2)
            })
    detail.sort(key=lambda z: (z.get('Año') or 0,z.get('Mes') or 0))
    return round(total,2), detail


def vacation_cycles(start: date, end: date, irows):
    """Ciclos por aniversario. Cada año completo se genera al aniversario siguiente; el último tramo queda proporcional."""
    if not start or not end or end < start:
        return []
    cycles=[]
    cursor=start
    n=1
    while True:
        try:
            anniv=date(cursor.year+1,cursor.month,cursor.day)
        except ValueError:
            anniv=date(cursor.year+1,cursor.month,28)
        cycle_end=anniv - pd.Timedelta(days=1)
        cycle_end = cycle_end.date() if hasattr(cycle_end,'date') else cycle_end
        if anniv <= end:
            base, det=sum_iess_base(irows,cursor,cycle_end)
            cycles.append({'Ciclo':n,'Desde':cursor,'Hasta':cycle_end,'Se genera el':anniv,'Tipo':'AÑO COMPLETO','Base IESS':base,'Vacación teórica':round(base/24,2),'Detalle':det})
            cursor=anniv
            n+=1
        else:
            base, det=sum_iess_base(irows,cursor,end)
            cycles.append({'Ciclo':n,'Desde':cursor,'Hasta':end,'Se genera el':None,'Tipo':'PROPORCIONAL AL CESE','Base IESS':base,'Vacación teórica':round(base/24,2),'Detalle':det})
            break
    return cycles


def legal_benefits_from_iess(start: date, end: date, irows, region='Costa / Insular', sbu=SBU_2026):
    """Calcula bases legales usando lo efectivamente reportado en IESS como fuente operativa de remuneración."""
    if not start or not end:
        return {}
    # Décimo tercero: período legal 1-dic a 30-nov; al cese se liquida lo acumulado del período corriente.
    if end.month == 12:
        d13_start=date(end.year,12,1)
    else:
        d13_start=date(end.year-1,12,1)
    d13_start=max(start,d13_start)
    d13_base,d13_detail=sum_iess_base(irows,d13_start,end)
    d13=round(d13_base/12,2)

    # Décimo cuarto: período legal regional; base de 360 días, no base IESS monetaria.
    p14_start,p14_end=dec14_period(end,region)
    a=max(start,p14_start); b=min(end,p14_end)
    d14_days=max(0, min(360, days360_us(a,b)+1)) if b>=a else 0
    d14=round(num(sbu)/360*d14_days,2)

    vac=vacation_cycles(start,end,irows)
    current_vac=next((c for c in reversed(vac) if c['Tipo']=='PROPORCIONAL AL CESE'), None)
    last_salary=latest_iess_salary(irows,end)
    years=completed_years(start,end)
    fund_reserve_from=None
    try:
        fund_reserve_from=date(start.year+1,start.month,start.day)
    except ValueError:
        fund_reserve_from=date(start.year+1,start.month,28)
    fr_base,fr_detail=(0.0,[])
    if end >= fund_reserve_from:
        fr_base,fr_detail=sum_iess_base(irows,fund_reserve_from,end)
    fund_reserve=round(fr_base/12,2) # 8.33% = 1/12
    return {
        'd13_period_start':d13_start,'d13_period_end':end,'d13_base_iess':d13_base,'d13_calc':d13,'d13_detail':d13_detail,
        'd14_period_start':a,'d14_period_end':b,'d14_days':d14_days,'d14_calc':d14,
        'vacation_cycles':vac,'vac_current_calc':round(current_vac['Vacación teórica'],2) if current_vac else 0.0,
        'vac_current_base':round(current_vac['Base IESS'],2) if current_vac else 0.0,
        'last_iess_salary':round(last_salary,2),'completed_years':years,
        'fund_reserve_start':fund_reserve_from,'fund_reserve_base_iess':fr_base,'fund_reserve_calc':fund_reserve,'fund_reserve_detail':fr_detail,
    }

# ===================== V8.3: CONTROL EXPLÍCITO DE PERIODOS DE BENEFICIOS =====================
def month_sequence(start: date, end: date):
    """Lista (año, mes) inclusive entre dos fechas."""
    if not start or not end or end < start:
        return []
    out=[]
    y,m=start.year,start.month
    while (y,m) <= (end.year,end.month):
        out.append((y,m))
        m += 1
        if m == 13:
            m=1; y+=1
    return out


def fmt_period_months(periods):
    return ', '.join(f'{m:02d}/{y}' for y,m in periods)


def rrhh_periods_from_item(item):
    out=[]
    for r in item.get('rows',[]):
        m=_month_num(r.get('Mes'))
        y=int(r.get('Año inferido') or (item.get('end').year if item.get('end') else date.today().year))
        if m:
            out.append((y,m))
    return list(dict.fromkeys(out))


def current_d13_window(start: date, end: date):
    """Ventana de décimo tercero vigente a la fecha de salida.
    Período legal: 1-dic a 30-nov; la fecha de ingreso recorta el inicio si es posterior.
    """
    if not start or not end or end < start:
        return None
    if end.month == 12:
        statutory_start=date(end.year,12,1); statutory_end=date(end.year+1,11,30)
    else:
        statutory_start=date(end.year-1,12,1); statutory_end=date(end.year,11,30)
    applied_start=max(start,statutory_start)
    applied_end=min(end,statutory_end)
    if applied_end < applied_start:
        return None
    full = applied_start == statutory_start and applied_end == statutory_end
    reasons=[]
    if applied_start > statutory_start: reasons.append('ingreso dentro del período')
    if applied_end < statutory_end: reasons.append('salida antes del cierre')
    kind='PERÍODO COMPLETO' if full else 'PROPORCIONAL'
    reason='; '.join(reasons) if reasons else 'período legal completo'
    return {'statutory_start':statutory_start,'statutory_end':statutory_end,
            'start':applied_start,'end':applied_end,'type':kind,'reason':reason,
            'months':month_sequence(applied_start,applied_end)}


def current_d14_window(start: date, end: date, region='Costa / Insular'):
    if not start or not end or end < start:
        return None
    statutory_start,statutory_end=dec14_period(end,region)
    applied_start=max(start,statutory_start); applied_end=min(end,statutory_end)
    if applied_end < applied_start:
        return None
    full = applied_start == statutory_start and applied_end == statutory_end
    reasons=[]
    if applied_start > statutory_start: reasons.append('ingreso dentro del período')
    if applied_end < statutory_end: reasons.append('salida antes del cierre')
    return {'statutory_start':statutory_start,'statutory_end':statutory_end,
            'start':applied_start,'end':applied_end,
            'type':'PERÍODO COMPLETO' if full else 'PROPORCIONAL',
            'reason':'; '.join(reasons) if reasons else 'período legal completo',
            'months':month_sequence(applied_start,applied_end)}


def current_vacation_window(start: date, end: date):
    """Ciclo vacacional que está vigente en la fecha de salida, basado en aniversario de ingreso."""
    if not start or not end or end < start:
        return None
    years=completed_years(start,end)
    try:
        cycle_start=date(start.year+years,start.month,start.day)
    except ValueError:
        cycle_start=date(start.year+years,start.month,28)
    # Si el aniversario coincide exactamente con la salida, el ciclo anterior quedó completo.
    if cycle_start == end and years > 0:
        try:
            cycle_start=date(start.year+years-1,start.month,start.day)
        except ValueError:
            cycle_start=date(start.year+years-1,start.month,28)
    try:
        next_anniv=date(cycle_start.year+1,cycle_start.month,cycle_start.day)
    except ValueError:
        next_anniv=date(cycle_start.year+1,cycle_start.month,28)
    full_end=next_anniv - pd.Timedelta(days=1)
    full_end=full_end.date() if hasattr(full_end,'date') else full_end
    applied_end=min(end,full_end)
    full = applied_end == full_end
    return {'start':cycle_start,'end':applied_end,'full_cycle_end':full_end,
            'type':'AÑO COMPLETO' if full else 'PROPORCIONAL AL CESE',
            'reason':'ciclo por aniversario de ingreso',
            'months':month_sequence(cycle_start,applied_end)}


def period_set_check(rrhh_periods, expected_periods):
    rr=set(rrhh_periods or []); ex=set(expected_periods or [])
    missing=sorted(ex-rr); extra=sorted(rr-ex)
    return {'missing':missing,'extra':extra,'ok':not missing and not extra}

def legal_benefits_from_iess(start: date, end: date, irows, region='Costa / Insular', sbu=SBU_2026):
    """V8.3: beneficios por sus períodos legales, recortados por ingreso y salida.

    - D13: 1-dic / 30-nov, proporcional si ingreso/salida recortan el período.
    - D14: período regional, proporcional por tiempo aplicable.
    - Vacaciones: ciclo individual por aniversario de la fecha de ingreso.
    Las bases monetarias conservan el IESS cuando reporta menos días y solo se recortan si exceden el máximo permitido por ingreso/salida.
    """
    if not start or not end or end < start:
        return {}

    w13=current_d13_window(start,end)
    if w13:
        d13_base,d13_detail=sum_iess_base(irows,w13['start'],w13['end'])
        d13=round(d13_base/12,2)
    else:
        d13_base,d13_detail,d13=0.0,[],0.0

    w14=current_d14_window(start,end,region)
    if w14:
        d14_days=max(0,min(360,days360_us(w14['start'],w14['end'])+1))
        d14=round(num(sbu)/360*d14_days,2)
    else:
        d14_days,d14=0,0.0

    all_vac=vacation_cycles(start,end,irows)
    wvac=current_vacation_window(start,end)
    if wvac:
        vac_base,vac_detail=sum_iess_base(irows,wvac['start'],wvac['end'])
        vac_current=round(vac_base/24,2)
    else:
        vac_base,vac_detail,vac_current=0.0,[],0.0

    last_salary=latest_iess_salary(irows,end)
    years=completed_years(start,end)
    try:
        fund_reserve_from=date(start.year+1,start.month,start.day)
    except ValueError:
        fund_reserve_from=date(start.year+1,start.month,28)
    fr_base,fr_detail=(0.0,[])
    if end >= fund_reserve_from:
        fr_base,fr_detail=sum_iess_base(irows,fund_reserve_from,end)
    fund_reserve=round(fr_base/12,2)

    return {
        'd13_period_start':w13['start'] if w13 else None,
        'd13_period_end':w13['end'] if w13 else None,
        'd13_statutory_start':w13['statutory_start'] if w13 else None,
        'd13_statutory_end':w13['statutory_end'] if w13 else None,
        'd13_type':w13['type'] if w13 else '', 'd13_reason':w13['reason'] if w13 else '',
        'd13_months':w13['months'] if w13 else [],
        'd13_base_iess':d13_base,'d13_calc':d13,'d13_detail':d13_detail,
        'd14_period_start':w14['start'] if w14 else None,
        'd14_period_end':w14['end'] if w14 else None,
        'd14_statutory_start':w14['statutory_start'] if w14 else None,
        'd14_statutory_end':w14['statutory_end'] if w14 else None,
        'd14_type':w14['type'] if w14 else '', 'd14_reason':w14['reason'] if w14 else '',
        'd14_months':w14['months'] if w14 else [], 'd14_days':d14_days,'d14_calc':d14,
        'vacation_cycles':all_vac,
        'vac_period_start':wvac['start'] if wvac else None,
        'vac_period_end':wvac['end'] if wvac else None,
        'vac_full_cycle_end':wvac['full_cycle_end'] if wvac else None,
        'vac_type':wvac['type'] if wvac else '', 'vac_reason':wvac['reason'] if wvac else '',
        'vac_months':wvac['months'] if wvac else [],
        'vac_current_calc':vac_current,'vac_current_base':vac_base,'vac_current_detail':vac_detail,
        'last_iess_salary':round(last_salary,2),'completed_years':years,
        'fund_reserve_start':fund_reserve_from,'fund_reserve_base_iess':fr_base,'fund_reserve_calc':fund_reserve,'fund_reserve_detail':fr_detail,
    }
